import time
import os.path
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import dataframes as ds
import out as ou
import all_roll as ar
import summary as sr
sheet_names = list(ds.room_dict.keys())

#print(sheet_names)
year = time.strftime("%y", time.localtime())
mon = time.strftime("%b", time.localtime())
date = time.strftime("%d", time.localtime())
excel_file_name = date+"_"+mon + "_" + year + "_" +"out_sheet" +".xlsx"
print(excel_file_name)
if not os.path.exists(excel_file_name):
    workbook = openpyxl.Workbook()
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    worksheet = workbook.create_sheet(date+"_"+mon+"_"+year+"_"+"out_sheet")
    black = "FF000000"
    yellow = "FFFFFF00"
    thin = Side(style="thin", color=black)
    alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_style = PatternFill(fill_type="solid", patternType='solid', start_color=yellow)
    worksheet.column_dimensions['B'].width = 20
    worksheet.merge_cells('A2:H2')
    worksheet['A2'].font = Font(size=23, name='Times New Roman')
    worksheet['A2'].alignment = alignment
    worksheet['A2'] = 'Mahaveer Institute of Science and Technology'
    worksheet['A2'].fill = fill_style
    worksheet['A2'].border = border_style
    worksheet.merge_cells('A3:H3')
    worksheet['A3'].font = Font(size=23, name='Times New Roman')
    worksheet['A3'].alignment = alignment
    worksheet['A3'] = 'NEW BLOCK'
    worksheet['A3'].fill = fill_style
    worksheet['A3'].border = border_style
    worksheet.merge_cells('A5:H5')
    worksheet['A5'].font = Font(size=23, name='Times New Roman')
    worksheet['A5'].alignment = alignment
    worksheet['A5'] = "JNTU External Examination-"+" "+mon+"- "+year
    worksheet['A5'].fill = fill_style
    worksheet['A5'].border = border_style
    worksheet.merge_cells('A4:H4')
    worksheet['A4'].font = Font(size=23, name='Times New Roman')
    worksheet['A4'].alignment = Alignment(horizontal='right')
    worksheet['A4']="Dt : "+date+"."+mon+"."+year
    worksheet['A4'].fill = fill_style
    worksheet['A4'].border = border_style
    worksheet.merge_cells('A6:H6')
    worksheet.merge_cells('A7:H7')
    worksheet['A6'].font = Font(size=23, name='Times New Roman')
    worksheet['A6'].alignment = alignment
    worksheet['A6'] = "Seating Arrangement"
    worksheet['A6'].fill = fill_style
    worksheet['A6'].border = border_style
    worksheet['A6'].font = Font(size=23, name='Times New Roman')
    worksheet['A6'].alignment = alignment
    worksheet['A6'] = "B.TECH "+str(int(ar.student_list['Year'].unique()))+"-"+str(
            int(ar.student_list['Semester'].unique()))+"  Reg/Supply Room Numbers"
    worksheet['A6'].fill = fill_style
    worksheet['A6'].border = border_style
    worksheet['A8']='Room No'
    worksheet['A8'].alignment = alignment
    worksheet['B8']='College Code'
    worksheet['B8'].alignment = alignment
    worksheet['C8']='Branch'
    worksheet['C8'].alignment = alignment
    worksheet.merge_cells('D8:G8')
    worksheet['D8']='H.T.Nos'
    worksheet['D8'].alignment = alignment
    worksheet['H8']='Total'
    worksheet['H8'].alignment = alignment
    #worksheet['I8']='Reg'
    #worksheet['I8'].alignment = alignment
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 7
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['H'].width = 7
    #worksheet.column_dimensions['I'].width = 7
    rows = range(9,33)

    # Merge the cells for each row
    for row in rows:
        worksheet.merge_cells(f'D{row}:G{row}')

    rows = range(8, 33)
    height = 75
    for row in rows:
        worksheet.row_dimensions[row].height = height
    current_row = 9
    for item in ou.data:
            # Write each key name to a new row in column C
        for i, keyname in enumerate(item['keynames']):
            worksheet.cell(row=current_row + i, column=3).value = keyname
            # Move the current row down by the number of keynames
        current_row += len(item['keynames'])
        start_row = 9
        start_column = 2

        for row_values in ou.code:
            for column_index, value in enumerate(row_values):
                value_str = ','.join(value)
                # Calculate the cell coordinates based on the row and column indexes
                cell = worksheet.cell(row=start_row + column_index, column=start_column)
                cell.value = value_str
            start_row += len(row_values)
    start_row = 9
    start_column = 8

    for i, d in enumerate(ou.for_room):
        # Iterate over each key in the dictionary
        for j, key in enumerate(d):
            # Get the length of the value for the current key
            value_len = len(d[key])
            worksheet.cell(row=start_row , column=start_column).value = value_len
            start_row+=1
    start_row = 9
    start_column = 4

    for i, d in enumerate(ou.for_room):
        # Iterate over each key in the dictionary
        for j, key in enumerate(d):
            # Get the length of the value for the current key
            values = d[key]
            values= ','.join(values)
            worksheet.cell(row=start_row, column=start_column).value = values
            start_row += 1
    row_num = 9
    for key, value in ou.merge_dict.items():
        if value==0:
            end_row=row_num
        else:
            end_row=row_num+value-1
        worksheet.merge_cells(start_row=row_num, start_column=1, end_row=end_row, end_column=1)
        worksheet.cell(row=row_num, column=1).value = key
        row_num += value
    thin = Side(border_style="thin", color="FF000000")
    for row in worksheet['A2':'H32']:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment=alignment#(wrap_text=True)
    for row in range(9, 33):
        cell = worksheet.cell(row=row, column=4)
        cell.alignment = Alignment(wrap_text=True)

    openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size=4, orientation='portrait')

else:
    workbook = openpyxl.load_workbook(excel_file_name)


workbook.save(excel_file_name)
workbook.close()
os.startfile(excel_file_name)
os.startfile('rooms.xlsx')