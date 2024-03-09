import openpyxl
import time
import os

year = time.strftime("%y", time.localtime())
mon = time.strftime("%b", time.localtime())
date = time.strftime("%d", time.localtime())
source_workbook = openpyxl.load_workbook(date+"_"+mon+"_"+year+"_"+'all_roll.xlsx')

source_worksheet = source_workbook.active
if not os.path.exists(date+'_'+mon+'_'+year+'_'+'halls_sheet.xlsx'):
    new_workbook = openpyxl.Workbook()

    new_worksheet = new_workbook.active

    new_worksheet['A1'] = 'Branch Name'
    new_worksheet['B1'] = 'Regulation'
    new_worksheet['C1'] = 'Total Students'

    total_students = {}

    for row in source_worksheet.iter_rows(min_row=2, values_only=True):
        branch_name = row[7]
        regulation = row[4]


        if (branch_name, regulation) not in total_students:
            total_students[(branch_name, regulation)] = 1

        else:
            total_students[(branch_name, regulation)] += 1

    row_num = 2
    for key, value in total_students.items():
        branch_name = key[0]
        regulation = key[1]
        new_worksheet.cell(row=row_num, column=1, value=branch_name)
        new_worksheet.cell(row=row_num, column=2, value=regulation)
        new_worksheet.cell(row=row_num, column=3, value=value)
        row_num += 1
    new_workbook.save(date + "_" + mon + "_" + year + "_" + 'summary.xlsx')
    new_workbook.close()
    os.startfile(date + "_" + mon + "_" + year + "_" + 'summary.xlsx')
