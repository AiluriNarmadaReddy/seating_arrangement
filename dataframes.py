import pandas as pd
import all_roll as ar
import openpyxl
import os.path
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import time


bname={
    '01': "CIV",
    '02': "EEE",
    '03': "MECH",
    '04': "ECE",
    '05': "CSE",
    '12': "IT",
    '21': "AERO",
    '66': "CSM",
    '27': "XYZ"
}
room_dict={}
year = time.strftime("%y", time.localtime())
mon = time.strftime("%b", time.localtime())
date = time.strftime("%d", time.localtime())
excel_file_name = date+"_"+ mon + "_" + year + ".xlsx"
rooms=pd.read_excel('rooms.xlsx')
i=0
j=0

for index, row in rooms.iterrows():
    room_number = row['Room No']
    rows = row['Rows']
    columns = row['Columns']
    if len(ar.student_list['Branch Name'].unique()) !=1:
        df = pd.DataFrame(index=range(1,rows+1),columns=range(1,columns+1))
        odd_cols = (columns + 1) // 2
        even_cols = columns - odd_cols
        for odd_col in range(1, odd_cols + 1):
            for odd_row in range(1, rows + 1):
                if i < len(ar.odd_list):
                    df.at[odd_row, odd_col * 2 - 1] = ar.odd_list[i]
                else:
                    df.at[odd_row, odd_col * 2 - 1] = None
                i += 1

        for even_col in range(1, even_cols + 1):
            for even_row in range(1, rows + 1):
                if j < len(ar.even_list):
                    df.at[even_row, even_col * 2] = ar.even_list[j]
                else:
                    df.at[even_row, even_col * 2 - 1] = None
                j += 1

        room_dict[room_number]=df
    else:
        k = 0
        students=ar.even_list+ar.odd_list
        df = pd.DataFrame(index=range(rows), columns=range(columns))
        for col in df.columns:
            for row in df.index:
                df.iat[row, col] = students[k]
                k = k + 1
        room_dict[room_number] = df
print(room_dict)
new_room_dict = {}
for room, df in room_dict.items():
    df.fillna(value='---', inplace=True)
    new_cols = {}
    for col in df.columns:
        hallticket_list = []
        for row in df.index:
            hallticket = df.iloc[row-1, col-1]
            if not pd.isna(hallticket):  # check if hallticket is not null
                hallticket_no = bname.get(hallticket[6:8], hallticket[6:8])
                hallticket_list.append(hallticket_no)
        header = list(set(hallticket_list))
        new_cols[col] = '  '.join(header)
    new_df = df.rename(columns=new_cols)
    new_room_dict[room] = new_df

room_dict = new_room_dict
print(room_dict)
if not os.path.exists(excel_file_name):	
    workbook = openpyxl.Workbook()
    for sheet_name, df in room_dict.items():
        sheet_name=str(sheet_name)
        worksheet = workbook.create_sheet(title=sheet_name)
        workbook = openpyxl.Workbook()
        for sheet_name, df in room_dict.items():
            sheet_name=str(sheet_name)
            worksheet = workbook.create_sheet(title=sheet_name)
            start_row = 7
            for col_idx, col_name in enumerate(df.columns, start=1):
                worksheet.cell(row=start_row, column=col_idx+1, value=col_name)

            for row_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
                for col_idx, value in enumerate(row, start=1):
                    worksheet.cell(row=row_idx, column=col_idx+1, value=value)
            branch_counts = df.applymap(lambda x: x[6:8]).stack().value_counts()            
            branch_counts.index = branch_counts.index.map(bname)
            print(branch_counts)
            worksheet = workbook[sheet_name]
            border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
            last_col = get_column_letter(df.shape[1]+1)
            black = "FF000000"
            yellow = "FFFFFF00"
            thin = Side(style="thin", color=black)
            alignment = Alignment(horizontal="center", vertical="center")
    
            fill_style = PatternFill(fill_type="solid", patternType='solid', start_color=yellow)
            merged_range = f'A2:{last_col}2'
            worksheet.merge_cells(merged_range)
            worksheet['A2'].font = Font(size=24, name='Times New Roman')
            worksheet['A2'].alignment = alignment
            worksheet['A2'] = 'Mahaveer Institute of Science and Technology'
            worksheet['A2'].fill = fill_style
            worksheet['A2'].border = border_style
            merged_range = f'A3:{last_col}3'
            worksheet.merge_cells(merged_range)
            worksheet['A3'].font = Font(size=24, name='Times New Roman')
            worksheet['A3'].alignment = alignment
            title_string = 'Seating Arrangement' + " " + str(int(ar.student_list['Year'].unique())) + "-" + str(
                int(ar.student_list['Semester'].unique())) + " " + 'External Examination  ' + mon + "  " + year
            worksheet['A3'] = title_string
            worksheet['A3'].fill = fill_style
            worksheet['A3'].border = border_style
    
            worksheet.merge_cells('A4:C4')
            worksheet['A4'].font = Font(size=24, name='Times New Roman')
            worksheet['A4'].alignment = alignment
            worksheet['A4'] = "ROOM NO:" + str(sheet_name)
            worksheet.merge_cells('D4:E4')
            merged_range = f'F4:{last_col}4'
            worksheet.merge_cells(merged_range)
            worksheet['F4'].font = Font(size=24, name='Times New Roman')
            worksheet['F4'].alignment = alignment
            worksheet['F4'] = "DT:" + date + "-" + mon + "-" + year
            merged_range = f'A5:{last_col}5'
            worksheet.merge_cells(merged_range)
            output_str = ''
            for index, value in branch_counts.items():
                output_str += str(index) + ' ' + str(value) + '\n'
            worksheet['A5'] = 'Strength:\n'+output_str
            worksheet['A5'].font = Font(size=24, name='Times New Roman')
            worksheet['A5'].alignment = Alignment(wrap_text=True)
    
    
            for column in ['A', 'B', 'C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q']:
                worksheet.column_dimensions[column].width = 15
    
    
            worksheet['A7'] = 'ENTRANCE'
            worksheet.column_dimensions['A'].auto_size = 'True'
            if len(ar.student_list['Branch Name'].unique()) != 1:
                num_cols = df.shape[1]
                start_col = 'B'
                end_col = chr(ord('A') + num_cols)
    
                # Merge cells and add row labels
                for i in range(0, num_cols, 2):
                    col_range = f'{chr(ord(start_col) + i)}6:{chr(ord(start_col) + i + 1)}6'
                    row_label = f'ROW {i // 2 + 1}'
                    worksheet.merge_cells(col_range)
                    worksheet[chr(ord(start_col) + i) + '6'] = row_label
                    worksheet[chr(ord(start_col) + i) + '6'].font = Font(size=24, name='Times New Roman')
                    worksheet[chr(ord(start_col) + i) + '6'].alignment = alignment
    
                # Unmerge last cell if there are an odd number of columns
                if num_cols % 2 == 1:
                    last_col_range = f'{chr(ord(end_col))}6:{chr(ord(end_col) + 1)}6'
                    worksheet.unmerge_cells(last_col_range)
            else:
                num_cols=df.shape[1]
                start_col='B'
                end_col=chr(ord('A') + num_cols)
                for i in range(0,num_cols,1):
                    col_range = f'{chr(ord(start_col) + i)}6:{chr(ord(start_col) + i + 1)}6'
                    row_label = f'ROW {i+ 1}'
                    worksheet[chr(ord(start_col) + i) + '6'] = row_label
                    worksheet[chr(ord(start_col) + i) + '6'].font = Font(size=24, name='Times New Roman')
                    worksheet[chr(ord(start_col) + i) + '6'].alignment = alignment
    
    
            for row in range(6,40):
                worksheet.row_dimensions[row].height=35
    
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row+1, min_col=0, max_col=worksheet.max_column+1):
                for cell in row:
                    if cell.value:
                        cell.border = border_style
            for merged_range in worksheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.min_col, \
                                                     merged_range.min_row, merged_range.max_col, merged_range.max_row
                for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                    for cell in row:
                            cell.border = border_style
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = alignment
            
            openpyxl.worksheet.worksheet.Worksheet.set_printer_settings(worksheet, paper_size=4, orientation='landscape')
    sheet_name = "Sheet"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
       
else:
    workbook = openpyxl.load_workbook(excel_file_name)
workbook.save(excel_file_name)
workbook.close()
os.startfile(excel_file_name)
