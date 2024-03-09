from tkinter import ttk , messagebox
import openpyxl
import tkinter as tk
import os.path
from openpyxl import Workbook, load_workbook

if os.path.isfile('rooms.xlsx'):
    workbook =load_workbook(filename='rooms.xlsx')
else:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Room No", "Rows", "Columns"])
    workbook.save("rooms.xlsx")
worksheet = workbook.active

halls = tk.Tk()
halls.geometry("700x500+10+10")
halls.title("Rooms")
tree = ttk.Treeview(halls)
 
tree["columns"] = ("Room No", "Rows", "Columns")

tree.column("#0", width=0, stretch=tk.NO)
tree.column("Room No", anchor=tk.CENTER, width=100)
tree.column("Rows", anchor=tk.CENTER, width=100)
tree.column("Columns", anchor=tk.CENTER, width=100)

tree.heading("#0", text="")
tree.heading("Room No", text="Room No")
tree.column("Room No", anchor=tk.CENTER, width=100)
tree.heading("Rows", text="Rows")
tree.column("Rows", anchor=tk.CENTER, width=100)
tree.heading("Columns", text="Columns")
tree.column("Columns", anchor=tk.CENTER, width=100)

style = ttk.Style()
style.configure("Treeview", rowheight=25, font=("Arial", 10))
style.configure("Treeview.Heading", font=("Arial", 10, "bold"))
style.layout("Custom.Treeview", [
    ("Custom.Treeview.treearea", {'sticky': 'nswe'}),
    ("Custom.Treeview.heading", {'sticky': 'we', 'children': [
        ("Custom.Treeview.heading.cell", {'sticky': 'we'})
    ]}),
    ("Custom.Treeview.row", {'sticky': 'nswe', 'children': [
        ("Custom.Treeview.cell", {'sticky': 'nswe', 'children': [
            ("Custom.Treeview.padding", {'sticky': 'nswe', 'children': [
                ("Custom.Treeview.text", {'sticky': 'w'})
            ]})
        ]}),
    ]}),
])

# Populate the Treeview with data from the Excel file
for row in worksheet.iter_rows(min_row=2,values_only=True):
    tree.insert("", tk.END, text="", values=row)

# Add the Treeview to the frame
tree.pack(expand=True, fill=tk.BOTH)

# Create a frame to hold the buttons
button_frame = ttk.Frame(halls)

button_frame.pack(side=tk.BOTTOM, pady=(10, 0))

# Create the buttons

add_button = ttk.Button(button_frame, text="Add", command=lambda:  add_button_click(tree))
edit_button = ttk.Button(button_frame, text="Edit", command=lambda: edit_button_click(tree))
delete_button = ttk.Button(button_frame, text="Delete", command=lambda: delete_row(tree))
exit_button=ttk.Button(button_frame,text='Exit',command=lambda:exit_row())
# Add the buttons to the frame
add_button.pack(side=tk.LEFT, padx=10)
edit_button.pack(side=tk.LEFT, padx=10)
delete_button.pack(side=tk.LEFT, padx=10)
exit_button.pack(side=tk.LEFT,padx=10)

window_open_popup = False
window_edit_popup=False

# Function to handle "Add" button click
def add_button_click(tree):
    global window_open_popup
    if not window_open_popup:
        open_add_row_popup(tree)
        window_open_popup = True
    else:
        messagebox.showinfo("Popup Already Open", "The 'Add Row' popup is already open.")

# Function to open "Add Row" popup
def open_add_row_popup(tree):
    global window_open_popup
    global popup1
    popup1 = tk.Toplevel()
    popup1.title("Add Row")
    popup1.geometry("+10+10")
    popup1.protocol("WM_DELETE_WINDOW", on_popup_close)

    room_no_var = tk.IntVar(value=0)
    rows_var = tk.IntVar(value=0)
    columns_var = tk.IntVar(value=0)
    room_no_label = ttk.Label(popup1, text="Room No.")
    room_no_entry = ttk.Entry(popup1, textvariable=room_no_var)
    rows_label = ttk.Label(popup1, text="Rows")
    rows_entry = ttk.Entry(popup1, textvariable=rows_var)
    columns_label = ttk.Label(popup1, text="Columns")
    columns_entry = ttk.Entry(popup1, textvariable=columns_var)

    room_no_label.pack()
    room_no_entry.pack()
    rows_label.pack()
    rows_entry.pack()
    columns_label.pack()
    columns_entry.pack()

    save_button = ttk.Button(popup1, text="Save", command=lambda: save_new_row(tree, room_no_var.get(), rows_var.get(), columns_var.get(), popup1))
    save_button.pack(pady=10)

# Function to handle popup close
def on_popup_close():
    global window_open_popup
    window_open_popup = False
    popup1.destroy()

# Function to save new row
def save_new_row(tree, room_no, rows, columns, popup):
    global window_open_popup  # Specify the global variable
    new_row = (room_no, rows, columns)
    worksheet.append(new_row)
    try:
        tree.insert("", tk.END, text="", values=new_row)
        workbook.save("rooms.xlsx")
    except PermissionError:
        messagebox.showerror("Error", "The file is already open. Please close it and try again.")
    popup1.destroy()
    window_open_popup = False

def edit_button_click(tree):
    global window_edit_popup
    if not window_edit_popup:
        open_edit_row_popup(tree)
        window_edit_popup = True
    else:
        messagebox.showinfo("Popup Already Open", "The 'edit Row' popup is already open.")
    
def open_edit_row_popup(tree):
    global window_edit_popup
    global popup2
    selected_row = tree.focus()
    if not selected_row:
        messagebox.showinfo('info',"Please click on a row you want to edit ")
    room_no, rows, columns = tree.item(selected_row)["values"]
    popup2 = tk.Toplevel()
    popup2.title("Edit Row")
    popup2.geometry("+10+10")
    popup2.protocol("WM_DELETE_WINDOW", on_popup2_close)
    room_no_var = tk.IntVar(value=room_no)
    rows_var = tk.IntVar(value=rows)
    columns_var = tk.IntVar(value=columns)
    room_no_label = ttk.Label(popup2, text="Room No.")
    room_no_entry = ttk.Entry(popup2, textvariable=room_no_var)
    rows_label = ttk.Label(popup2, text="Rows")
    rows_entry = ttk.Entry(popup2, textvariable=rows_var)
    columns_label = ttk.Label(popup2, text="Columns")
    columns_entry = ttk.Entry(popup2, textvariable=columns_var)
    room_no_label.pack()
    room_no_entry.pack()
    rows_label.pack()
    rows_entry.pack()
    columns_label.pack()
    columns_entry.pack()
    save_button = ttk.Button(popup2, text="Save", command=lambda: save_row(tree, selected_row, room_no_var.get(), rows_var.get(), columns_var.get(), popup2))
    save_button.pack(pady=10)
def on_popup2_close():
    global window_edit_popup
    window_edit_popup = False
    popup2.destroy()
def save_row(tree, selected_row, room_no, rows, columns, popup2):
    index = int(tree.index(selected_row))

    try:
        tree.item(selected_row, values=(room_no, rows, columns))
        worksheet.cell(row=index + 2, column=1, value=room_no)
        worksheet.cell(row=index + 2, column=2, value=rows)
        worksheet.cell(row=index + 2, column=3, value=columns)
        workbook.save("rooms.xlsx")
    except PermissionError:
        messagebox.showerror("Error", "The file is already open. Please close it and try again.")
        workbook.save("rooms.xlsx")
    global window_edit_popup
    window_edit_popup = False
    popup2.destroy()

def delete_row(tree):
    selected_row = tree.focus()
    if not selected_row:
        messagebox.showinfo('info',"Please click on a row you want to delete ")
    index =  int(tree.index(selected_row))

    try:
        tree.delete(selected_row)
        worksheet.delete_rows(index + 2)
        workbook.save("rooms.xlsx")
    except PermissionError:
        messagebox.showerror("Error", "The file is already open. Please close it and try again.")
        workbook.save("rooms.xlsx")
def exit_row():
    global window_open_rooms
    window_open_rooms = False
    halls.destroy()
halls.protocol("WM_DELETE_WINDOW", exit_row)

halls.mainloop()
